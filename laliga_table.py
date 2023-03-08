from selenium.webdriver.chrome.service import Service
from selenium import webdriver
from selenium.webdriver.common.by import By
import pandas as pd
import time
from selenium.webdriver.chrome.options import Options
import openpyxl
def extract():
        global df
        options=Options()
        options.add_argument("headless")
        driver = webdriver.Chrome()

        url='https://www.laliga.com/en-GB/laliga-santander/standing'
        driver.get(url)
        driver.maximize_window()
        time.sleep(20)
        print("Collecting laliga table data")
        pos=[]
        tname=[]
        points=[]
        pld=[]
        w=[]
        d=[]
        l=[]
        gf=[]
        ga=[]
        gd=[]
        tpath='/html/body/div[1]/div[6]/div[2]/div/div[3]/div[1]/div[1]/div/div[3]/div[1]/div/div/div[2]/div[2]/p'
        tname.append(driver.find_element(By.XPATH,tpath).text)
        pos.append(1)
        ppath='/html/body/div[1]/div[6]/div[2]/div/div[3]/div[1]/div[1]/div/div[3]/div[1]/div/div/div[3]/p'
        points.append(driver.find_element(By.XPATH,ppath).text)
        mpath='/html/body/div[1]/div[6]/div[2]/div/div[3]/div[1]/div[1]/div/div[4]/div[1]/div/div/div[4]/p'
        pld.append(driver.find_element(By.XPATH,mpath).text)
        wpath='/html/body/div[1]/div[6]/div[2]/div/div[3]/div[1]/div[1]/div/div[3]/div[1]/div/div/div[5]/p'
        w.append(driver.find_element(By.XPATH,wpath).text)
        dpath='/html/body/div[1]/div[6]/div[2]/div/div[3]/div[1]/div[1]/div/div[3]/div[1]/div/div/div[6]/p'
        d.append(driver.find_element(By.XPATH,dpath).text)
        lpath='/html/body/div[1]/div[6]/div[2]/div/div[3]/div[1]/div[1]/div/div[3]/div[1]/div/div/div[7]/p'
        l.append(driver.find_element(By.XPATH,lpath).text)
        
        for i in range(4,23):
                pos.append(i-2)
                tpath='/html/body/div[1]/div[6]/div[2]/div/div[3]/div[1]/div[1]/div/div[{}]/div[1]/div/div/div[2]/div[2]/p'.format(i)
                ppath='/html/body/div[1]/div[6]/div[2]/div/div[3]/div[1]/div[1]/div/div[{}]/div[1]/div/div/div[3]/p'.format(i)
                mpath='/html/body/div[1]/div[6]/div[2]/div/div[3]/div[1]/div[1]/div/div[{}]/div[1]/div/div/div[4]/p'.format(i)
                wpath='/html/body/div[1]/div[6]/div[2]/div/div[3]/div[1]/div[1]/div/div[{}]/div[1]/div/div/div[5]/p'.format(i)
                dpath='/html/body/div[1]/div[6]/div[2]/div/div[3]/div[1]/div[1]/div/div[{}]/div[1]/div/div/div[6]/p'.format(i)
                lpath='/html/body/div[1]/div[6]/div[2]/div/div[3]/div[1]/div[1]/div/div[{}]/div[1]/div/div/div[7]/p'.format(i)
                tname.append(driver.find_element(By.XPATH,tpath).text)
                points.append(driver.find_element(By.XPATH,ppath).text)
                pld.append(driver.find_element(By.XPATH,mpath).text)
                w.append(driver.find_element(By.XPATH,wpath).text)
                d.append(driver.find_element(By.XPATH,dpath).text)
                l.append(driver.find_element(By.XPATH,lpath).text)
                
                
                
                
                
        table=dict(Position =pos, Team =tname,Points=points, Matches_played=pld,Win=w,Draw=d,Lost=l)
        driver.close()
        df = pd.DataFrame.from_dict(table)
        
                
def load():
        print("Loading data...")
        wb=openpyxl.load_workbook('C:/Users/vggan/Desktop/Barca_project/barca_stats.xlsx')
        sheet = wb.sheetnames
        if "TABLE" in sheet:
                wb.remove(wb.get_sheet_by_name("TABLE"))
        wb.save('C:/Users/vggan/Desktop/Barca_project/barca_stats.xlsx')

                
        with pd.ExcelWriter('C:/Users/vggan/Desktop/Barca_project/barca_stats.xlsx') as writer:
                writer.book = openpyxl.load_workbook('C:/Users/vggan/Desktop/Barca_project/barca_stats.xlsx')
                df.to_excel(writer, sheet_name='TABLE',index=False)
       

if __name__ == '__main__':
        extract()
        load()

                
